import os
import openpyxl
import pandas as pd 

from django.contrib import messages
from django.shortcuts import render,redirect
from django.conf import settings

from .models import Patronal, Gasto, Entidad, Motivo

# Create your views here.
def create_folder_if_not_exists(folder_path):
    # Check if the folder path exists
    if not os.path.exists(folder_path):
        try:
            # Create the folder if it doesn't exist
            os.makedirs(folder_path)
            print(f"Folder '{folder_path}' created successfully.")
        except OSError as e:
            print(f"Error: Failed to create folder '{folder_path}'. Reason: {e}")
    else:
        print(f"Folder '{folder_path}' already exists.")

def save_uploaded_file(uploaded_file, path):
    with open(path, 'wb') as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)

    return path

def process_entidad_row(row):
    # Assuming the values are in the following order:
    # NIT, idTipoGasto (foreign keys), concepto, razonEntidad, rubro, tipoCuentaPagar, codigo

    # Get the Gasto instance that matches the value of idTipoGasto in the xlsx file
    try:
        gasto_instance = Gasto.objects.get(tipo=row[1])
    except Gasto.DoesNotExist:
        print(f"Error: No Gasto instance found with type '{row[1]}'")
        return
    except IndexError:
        print("Error: Column 'idTipoGasto' not found in the file")
        return

    # Create an Entidad instance and save it to the database
    entidad = Entidad(
        NIT=row[0],
        idTipoGasto=gasto_instance,
        concepto=row[2],
        razonEntidad=row[3],
        rubro=row[4],
        tipoCuentaPagar=row[5],
        codigo=row[6]
    )
    entidad.save()

def upload_data_entidades(request):
    if request.method == 'POST' and request.FILES.get('formFile'):
        upload_file = request.FILES['formFile']
        new_file_name = 'datos entidades.xlsx'
        path = os.path.join(settings.MEDIA_ROOT, new_file_name)
        save_path = save_uploaded_file(upload_file, path)
        show_alert = False

        try:
            workbook = openpyxl.load_workbook(save_path)
            worksheet = workbook.active
            last_row = worksheet.max_row

            context = {
                'show_alert': True,
                'alert_type':"success",
                'message':'El archivo se ha procesado correctamente.',
            }

            for row in worksheet.iter_rows(min_row=2, max_row=last_row, values_only=True):
                process_entidad_row(row)

            # You can perform more operations or redirect to a success page here
            return render(request, 'load_data_entidades.html', context)

        except Exception as e:
            context = {
                'show_alert': True,
                'alert_type':"danger",
                'message':'Ocurrió un error al procesar el archivo.',
            }

            # Error handling if something goes wrong while processing the file
            print(f"Error processing the file: {str(e)}")

            # You can add an error message in the response or redirect to an error page
            return render(request, 'load_data_entidades.html', context)

    return render(request, 'load_data_entidades.html')

def upload_documents( request ):
    if request.method == 'POST':
        # Acceder a los datos del formulario que se envían a través del método POST
        selected_year = request.POST.get('selectYear')
        selected_month = request.POST.get('selectMonth')

        # Definir la lista de diccionarios para los archivos
        filesDict = [
            {
                'nombreFormulario': 'planilla',
                'nuevoNombreArchivo': f'Planilla Detallada {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Planilla Detallada {selected_year}-{selected_month} converted.csv',
            },
            # {
            #     'nombreFormulario': 'patronalesTemporales',
            #     'nuevoNombreArchivo': f'Patronales Temporales {selected_year}-{selected_month}.xlsx',
            #     'nuevoNombreArchivoCSV': f'Planilla Detallada {selected_year}-{selected_month} converted.csv',
            # },
            # {
            #     'nombreFormulario': 'patronalesPermanentes',
            #     'nuevoNombreArchivo': f'Patronales Permanentes {selected_year}-{selected_month}.xlsx',
            #     'nuevoNombreArchivoCSV': f'Planilla Detallada {selected_year}-{selected_month} converted.csv',
            # },
            {
                'nombreFormulario': 'deduc2',
                'nuevoNombreArchivo': f'Deducibles Unidad 2 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 2 {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'deduc8',
                'nuevoNombreArchivo': f'Deducibles Unidad 8 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 8 {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'deduc9',
                'nuevoNombreArchivo': f'Deducibles Unidad 9 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 9 {selected_year}-{selected_month} converted.csv',
            },

        ]

        folder_path_xlsx = os.path.join(settings.MEDIA_ROOT, 'xlsx', selected_year, selected_month)
        folder_path_csv = os.path.join(settings.MEDIA_ROOT, 'csv', selected_year, selected_month)
        
        create_folder_if_not_exists(folder_path_xlsx)
        create_folder_if_not_exists(folder_path_csv)

        for file_info in filesDict:
            form_name = file_info['nombreFormulario']
            new_filename = file_info['nuevoNombreArchivo']
            new_filename_csv = file_info['nuevoNombreArchivoCSV']
            
            file = request.FILES.get(form_name)
            path_file_xlsx = os.path.join(folder_path_xlsx, new_filename)
            path_file_csv = os.path.join(folder_path_csv, new_filename_csv)

            save_uploaded_file(file, path_file_xlsx)
            converter_xlsx_to_csv(path_file_xlsx,path_file_csv)
    
    return render( request, 'load_documents.html')

def converter_xlsx_to_csv( folder_path_xlsx, folder_path_csv):
    try:
        # Cargar el archivo XLSX usando pandas
        dataframe = pd.read_excel(folder_path_xlsx)

        # Guardar el dataframe como archivo CSV
        dataframe.to_csv(folder_path_csv, index=False)
        print("Archivo convertido exitosamente a CSV.")
    except Exception as e:
        print("Error al convertir el archivo:", e)