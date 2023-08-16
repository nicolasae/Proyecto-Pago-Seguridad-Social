import os
import openpyxl
from datetime import datetime

from django.contrib import messages
from django.shortcuts import render,redirect
from django.conf import settings

from .process_files import *
from .process_data import *
from .process_data_entidad import *
from .process_data_planilla import *
from .process_data_patronales import * 
from .process_data_deducciones import * 


def delete_records_by_date(model, date_field_name, target_date):
    try:
        records_to_delete = model.objects.filter(**{date_field_name: target_date})
        records_to_delete.delete()
        print(f"Registros borrados en {model._meta.verbose_name_plural}: {len(records_to_delete)}")
    except Exception as e:
        print(f"Error al borrar registros: {str(e)}")

def upload_data_entidades(request):
    if request.method == 'POST':
        if request.FILES.get('formFile'):
            upload_file = request.FILES['formFile']
            new_file_name = 'datos_entidades.xlsx'
            path = os.path.join(settings.MEDIA_ROOT, new_file_name)
            save_path = save_uploaded_file(upload_file, path)
        
            try:
                workbook = openpyxl.load_workbook(save_path)
                worksheet = workbook.active
                last_row = worksheet.max_row

                for row in worksheet.iter_rows(min_row=2, max_row=last_row, values_only=True):
                    process_entidad_row(row)

                # Success message
                messages.success(request, "El archivo se ha procesado correctamente.")
            except Exception as e:
                # Error message
                messages.error(request, f"Ocurrió un error al procesar el archivo: {str(e)}, Intentelo de nuevo más tarde")
                print(f"Error processing the file: {str(e)}")

            return render(request, 'load_data_entidades.html')
        else:
            messages.error(request, "Por favor, seleccione un archivo para subir.")
            return render(request, 'load_data_entidades.html')

    return render(request, 'load_data_entidades.html')

def process_document_upload_files(request, data):
    # Lista para almacenar los nombres de los formularios que no se proporcionaron.
        missing_files = []

        # Loop through each file type and process the uploaded files.
        for file_info in data['filesDict']:
            form_name = file_info['nombreFormulario']
            new_filename = file_info['nuevoNombreArchivo']
            new_filename_csv = file_info['nuevoNombreArchivoCSV']
            
            # Get the uploaded file for the current file type.
            file = request.FILES.get(form_name)

            # Check if a file was provided for the current form_name before proceeding.
            if file:
                # Set the file paths for the xlsx and csv files.
                path_file_xlsx = os.path.join(data['folder_path_xlsx'], new_filename)
                path_file_csv = os.path.join(data['folder_path_csv'], new_filename_csv)

                # Save the uploaded file to the xlsx folder.
                save_uploaded_file(file, path_file_xlsx)

                # Convert the xlsx file to csv format and save it in the csv folder.           
                converter_xlsx_to_csv(path_file_xlsx, path_file_csv)

                # Clean up the csv file by removing any empty rows.
                clean_empty_rows_csv(path_file_csv)
                
                # Based on the type of form, extract data from the processed csv file and save it.
                if form_name == 'planilla':
                    extract_data_for_planilla(path_file_csv, data['selected_year'], data['selected_month'])
                if form_name == 'patronalesTemporales':
                    tipo_patronal = 'temporal'
                    extract_data_patronales(path_file_csv, data['selected_year'], data['selected_month'], tipo_patronal)
                if form_name == 'patronalesPermanentes':
                    tipo_patronal = 'permanente'
                    extract_data_patronales(path_file_csv, data['selected_year'], data['selected_month'], tipo_patronal)
                if form_name == 'deduc2':
                    unidad = 2
                    extract_data_deducciones(path_file_csv, data['selected_year'], data['selected_month'],unidad)
                if form_name == 'deduc8':
                    unidad = 8
                    extract_data_deducciones(path_file_csv, data['selected_year'], data['selected_month'],unidad)
                if form_name == 'deduc9':
                    unidad = 9
                    extract_data_deducciones(path_file_csv, data['selected_year'], data['selected_month'],unidad)
            else:
                equivalences_names_files = {
                    'planilla':'Planilla Detallada',
                    'patronalesTemporales':'Patronales de Planta Temporal',
                    'patronalesPermanentes':'Patronales de Planta Permanente',
                    'deduc2':'Deducciones Unidad 2',
                    'deduc8':'Deducciones Unidad 8',
                    'deduc9':'Deducciones Unidad 9'                    
                }
                # If the file was not provided, add the name of the form to the list of missing files.
                missing_files.append(equivalences_names_files[form_name])

        return missing_files

def upload_documents( request ):    
    if request.method == 'POST':
        # Access the data from the form submitted via the POST method.
        selected_year = request.POST.get('selectYear')
        selected_month = request.POST.get('selectMonth')

        # Define a list of dictionaries representing different file types and their new names.
        filesDict = [
            {
                'nombreFormulario': 'planilla',
                'nuevoNombreArchivo': f'Planilla Detallada {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Planilla Detallada {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'patronalesTemporales',
                'nuevoNombreArchivo': f'Patronales Temporales {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Patronales Temporales {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'patronalesPermanentes',
                'nuevoNombreArchivo': f'Patronales Permanentes {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Patronales Permanentes {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'deduc2',
                'nuevoNombreArchivo': f'Deducibles Unidad 2 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 2 {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'deduc8',
                'nuevoNombreArchivo': f'Deducibles Unidad 8 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 8 {selected_year}-{selected_month} converted.csv',
            },
            {
                'nombreFormulario': 'deduc9',
                'nuevoNombreArchivo': f'Deducibles Unidad 9 {selected_year}-{selected_month}.xlsx',
                'nuevoNombreArchivoCSV': f'Deducibles Unidad 9 {selected_year}-{selected_month} converted.csv',
            },
        ]

        # Define the folder paths for saving the uploaded files in xlsx and csv formats.
        folder_path_xlsx = os.path.join(settings.MEDIA_ROOT, 'xlsx', selected_year, selected_month)
        folder_path_csv = os.path.join(settings.MEDIA_ROOT, 'csv', selected_year, selected_month)
        
        # Ensure that the folders for saving the files exist; if not, create them.
        create_folder_if_not_exists(folder_path_xlsx)
        create_folder_if_not_exists(folder_path_csv)
        
        data = {
            'filesDict':filesDict, 
            'folder_path_xlsx':folder_path_xlsx,
            'folder_path_csv':folder_path_csv,
            'selected_year':selected_year,
            'selected_month':selected_month
        }

        date = f"{data['selected_year']}/{data['selected_month']}"
        delete_records_by_date(valoresPlanilla, 'numeroPlanilla__fecha', date)
        delete_records_by_date(valoresPatron, 'fecha', date)
        delete_records_by_date(valoresEmpleado, 'fecha', date)


        missing_files = process_document_upload_files(request, data)
        # Determine if all files were provided or if some are missing.
        show_alert = len(missing_files) == 0
        alert_type = "success" if show_alert else "danger"
        message = 'Los archivos se han procesado correctamente.' if show_alert else 'Faltan archivos por subir.'

    else:
        # If it is not a POST request, initialize variables so that the alert is not shown.
        missing_files = []
        show_alert = False
        alert_type = "success"  # Puedes establecerlo a "danger" si deseas un mensaje de error por defecto.
        message = ''

    # Dictionary with the context to pass to the 'load_documents.html' template.
    context = {
        'missing_files': missing_files,
        'show_alert': show_alert,
        'alert_type': alert_type,
        'message': message,
    }

    # Render the 'load_documents.html' template with the provided context.
    return render(request, 'load_documents.html', context)

# Render views
def render_list_entidades (request):
    entidades = Entidad.objects.all()
    return render(request, 'list_entidades.html', {'entidades':entidades})

def render_list_plantillas(request):
    return render (request, 'list_plantillas.html')