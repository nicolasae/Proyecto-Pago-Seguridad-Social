import os
import openpyxl

from django.shortcuts import render
from django.conf import settings

from .models import Gasto, Entidad

def lista_gastos(request):
    # Get all Gasto objects from the database
    gastos = Gasto.objects.all()
    return render(request, 'entidad/lista_gastos.html', {'gastos': gastos})


def cargar_datos_entidades(request):
    if request.method == 'POST' and request.FILES['formFile']:
        archivo_adjunto = request.FILES['formFile']
        
        # Specific name for the file you want to save
        new_file_name = "datos_entidades.xlsx"
        save_path = os.path.join(settings.STATIC_ROOT, new_file_name)

        # Save the xlsx file to the desired directory (in the "media" folder)
        with open(save_path, 'wb') as destination:
            for chunk in archivo_adjunto.chunks():
                destination.write(chunk)

        # Process the attached file here and save the data to the database
        try:
            workbook = openpyxl.load_workbook(save_path)
            worksheet = workbook.active
            last_row = worksheet.max_row

            for row in worksheet.iter_rows(min_row=2, max_row=last_row, values_only=True):
                # 'row' contains the values of each row from the xlsx file
                # Assuming the values are in the following order:
                # NIT, idTipoGasto (foreign keys), concepto, razonEntidad, rubro, tipoCuentaPagar, codigo

                # Get the Gasto instance that matches the value of idTipoGasto in the xlsx file
                try:
                    gasto_instance = Gasto.objects.get(tipo=row[1])
                except Gasto.DoesNotExist:
                    print(f"Error: No Gasto instance found with type '{row[1]}'")
                    continue
                except IndexError:
                    print("Error: Column 'idTipoGasto' not found in the file")
                    break
                
                # Create an Entidad instance and save it to the database
                entidad = Entidad(
                    NIT=row[0],
                    idTipoGasto=gasto_instance,
                    concepto=row[2],
                    razonEntidad=row[3],
                    rubro=row[4],
                    tipoCuentaPagar=row[5],
                    codigo=row[6]
                )
                entidad.save()

            # You can perform more operations or redirect to a success page here
            # return render(request, 'archivo_cargado.html')

        except Exception as e:
            # Error handling if something goes wrong while processing the file
            print(f"Error processing the file: {str(e)}")
            # You can add an error message in the response or redirect to an error page
            # return render(request, 'error.html')

    return render(request, 'entidad/cargar_datos_entidades.html')
