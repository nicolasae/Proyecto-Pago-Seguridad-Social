import openpyxl

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path

from document_upload.models import *

def download_view(request):
    if request.method == 'POST':
        selected_year = request.POST.get('selectYear')
        selected_month = request.POST.get('selectMonth')

        if 'btn_resumen_planilla' in request.POST:
            # Acción para generar el resumen de la planilla
            return create_report(selected_year,selected_month)
        
        if 'btn_resumen_patronales_temporales' in request.POST:
            # Aquí puedes realizar la acción para generar el resumen de las patronales temporales
            print("Botón Resumen Patronales Temporales presionado")
        
        if 'btn_resumen_patronales_permanentes' in request.POST:
            # Aquí puedes realizar la acción para generar el resumen de las patronales permanentes
            print("Botón Resumen Patronales Permanentes presionado")       
    return redirect('reportes')

def create_report(year,month):
    date = year + '/' + month
    # Obtener los objetos de infoPlanilla filtrados por año y mes
    info_planilla = infoPlanilla.objects.filter(periodo=date)

    # Crear un archivo de Excel para los datos de la planilla
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Info Planilla"

    # Mapeo de campos del modelo con los encabezados en el archivo Excel
    field_mapping = {
        'razonSocial': 'RAZON SOCIAL',
        'periodo':'PERIODO',
        'identificacion': 'IDENTIFICACION',
        'codigoDependenciaSucursal': 'COD. DEPENDENCIA O SUCURSAL',
        'nomDependenciaSucursal': 'NOM. DEPENDENCIA O SUCURSAL',
        'fechaReporte': 'FECHA GENERACION REPORTE',
        'fechaLimitePago': 'FECHA LIMITE DE PAGO',
        'periodoPension': 'PERIODO PENSION',
        'periodoSalud': 'PERIODO SALUD',
        'numeroPlanilla': 'NUMERO PLANILLA',
        'totalCotizantes': 'TOTAL COTIZANTES',
        'PIN': 'REFERENCIA DE PAGO (PIN)',
        'tipoPlanilla': 'TIPO DE PLANILLA',
    }

    # # Escribir los encabezados en la columna A
    for row_num, header in enumerate(field_mapping.values(), start=1):
        sheet.cell(row=row_num, column=1, value=header)

    row_num = 1
    for obj in info_planilla:
        for key in field_mapping:
            attribute_name = key  # Get the attribute name from the field_mapping
            value = getattr(obj, attribute_name, '')  # Get the attribute value using getattr
            sheet.cell(row = row_num, column = 2, value = value)
            row_num += 1

    # Crear la respuesta HTTP y devolver el archivo para su descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Planilla_Detallada-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

# def create_report_patronales_temporales(request):
#     print('entro a temporales')
#     return render(request, 'reports.html')

