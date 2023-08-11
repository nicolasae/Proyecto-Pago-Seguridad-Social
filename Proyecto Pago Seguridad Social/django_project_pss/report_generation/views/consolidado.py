import openpyxl 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.db.models import Sum,Case, CharField, Value, When
from django.core.exceptions import ObjectDoesNotExist
from document_upload.models import *
from .functions import *
from .constants import *

def convert_empty_to_zero(value):
    return value if value != None else 0

def generate_excel_report_consolidado(data, year, month):
    source_file  = 'media/plantillas/Consolidado.xlsx'
    # Load the existing excel file
    source_workbook = load_workbook(source_file)

    # Get the first sheet of the existing excel file
    source_sheet = source_workbook.active

    # Create a new Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active

    cells_range_entidad = "A1:C1"
    cells_range_empleado = "D1:G1"
    cells_range_patron = "H1:N1"

    # # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet1)
    
    sheet1.merge_cells(cells_range_entidad)
    sheet1.merge_cells(cells_range_empleado)
    sheet1.merge_cells(cells_range_patron)

    merged_ranges = [cells_range_entidad, cells_range_empleado, cells_range_patron]

    # Set alignment for merged cells
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(":")
        start_row, start_column = openpyxl.utils.coordinate_to_tuple(start_cell)
        end_row, end_column = openpyxl.utils.coordinate_to_tuple(end_cell)
        
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = sheet1.cell(row=row, column=column)
                cell.alignment = center_alignment
                cell.font = bold_font

    # Assign the desired name to the sheet
    sheet_name = f"Datos-{year}-{month}"
    sheet1.title = sheet_name

    # Add additional information to the sheet
    save_data(sheet1, data)

    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Consolidado-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

def get_data_values(date):
    try:
        empleados = valoresEmpleado.objects.filter(fecha=date).order_by('NIT__razonEntidad')
        empleados_ordenados = sorted(
            empleados, key=lambda empleado: PERSONALIZED_ORDER.index(empleado.NIT.razonEntidad)
        )

        patrones = valoresPatron.objects.filter(fecha=date).order_by('NIT__razonEntidad')
        patrones_ordenados = sorted(
            patrones, key=lambda patron: PERSONALIZED_ORDER.index(patron.NIT.razonEntidad)
        )

        # Create a dictionary to map entity names to indices
        entidad_to_index = {entidad: index for index, entidad in enumerate(PERSONALIZED_ORDER)}
        
        # Get the filtered objects and assign them order values
        valores_planilla_ordenados = valoresPlanilla.objects.filter(
            numeroPlanilla__fecha=date
        ).annotate(
            entidad_order=Case(
                *[When(codigoEntidad__concepto=entidad, then=Value(index)) for entidad, index in entidad_to_index.items()],
                default=Value(len(PERSONALIZED_ORDER)), output_field=CharField()
            )
        ).order_by('entidad_order')

        # Create a dictionary to store the information grouped by NIT
        data = {
            'Valores patron':patrones_ordenados,
            'Valores empleado':empleados_ordenados,
            'Valores planilla': valores_planilla_ordenados
        }
    
        return data
    
    except ObjectDoesNotExist:
        return None  # Return None if there's no data for the given date or any other related error
    except Exception as e:
        # Handle other exceptions as needed, such as database errors, unexpected behavior, etc.
        print(f"An error occurred: {e}")
        return None

def process_data_empleados(data):
    # Create a dictionary to store the information grouped by NIT
    data_dict = {}

    for motivo in data:
        NIT = motivo.NIT.NIT
        rubro = motivo.NIT.rubro
        concepto = motivo.NIT.concepto
        unidad = motivo.unidad

        # # Check if an entry for the NIT already exists in the dictionary
        if NIT not in data_dict:
            data_dict[NIT] = {
                'RUBRO': rubro,
                'CONCEPTO': concepto,
                'UNIDAD 2': 0,
                'UNIDAD 8': 0,
                'UNIDAD 9': 0,
                'TOTAL': 0
            }

        # # Add the information to the corresponding entry of type_employer
        data_dict[NIT]['TOTAL'] += motivo.saldo

        if unidad == 2:
            data_dict[NIT]['UNIDAD 2'] += motivo.saldo
        elif unidad == 8:
            data_dict[NIT]['UNIDAD 8'] += motivo.saldo
        elif unidad == 9:
            data_dict[NIT]['UNIDAD 9'] += motivo.saldo
    
    return data_dict

def process_data_patron(data):
    # Create a dictionary to store the information grouped by NIT
    data_dict = {}

    for motivo in data:
        NIT = motivo.NIT.NIT
        tipo_patronal = motivo.tipoPatronal.tipo
        rubro = motivo.NIT.rubro
        concepto = motivo.NIT.concepto
        codigo_descuento = motivo.NIT.codigoDescuento
        total = motivo.total

        # Check if an entry for the NIT already exists in the dictionary
        if NIT not in data_dict:
            data_dict[NIT] = {
                'RUBRO': rubro,
                'CONCEPTO': concepto,
                'CODIGO DEL CONCEPTO DE DESCUENTO': codigo_descuento,
                'TEMPORAL UN 2': 0,
                'TEMPORAL UN 8': 0,
                'TEMPORAL UN 9': 0,
                'PERMANENTE UN 2': 0,
                'PERMANENTE UN 8': 0,
                'PERMANENTE UN 9': 0,
                'TOTAL': 0
            }

        # Add the information to the corresponding entry of type_employer
        if tipo_patronal == 'temporal':
            data_dict[NIT]['TEMPORAL UN 2'] += motivo.unidad2
            data_dict[NIT]['TEMPORAL UN 8'] += motivo.unidad8
            data_dict[NIT]['TEMPORAL UN 9'] += motivo.unidad9
        elif tipo_patronal == 'permanente':
            data_dict[NIT]['PERMANENTE UN 2'] += motivo.unidad2
            data_dict[NIT]['PERMANENTE UN 8'] += motivo.unidad8
            data_dict[NIT]['PERMANENTE UN 9'] += motivo.unidad9

        # We also store the information in the entry of the employer type
        data_dict[NIT][tipo_patronal] = data_dict[NIT].get(tipo_patronal, 0) + motivo.unidad2

        data_dict[NIT]['TOTAL'] += total

    return data_dict

def process_data_planilla(data):
    # Create a dictionary to store the information grouped by NIT
    data_dict = {}
    for planilla in data:

        NIT = planilla.NIT
        concepto = planilla.codigoEntidad.concepto
        valor = planilla.valorPagar

        # Check if an entry for the NIT already exists in the dictionary
        if NIT not in data_dict:
            data_dict[NIT] = {
                'RAZON':concepto,
                'VALOR': valor,                
            }
        else:
            # If an entry for the NIT already exists, accumulate the value
            data_dict[NIT]['VALOR'] += valor
    
    return data_dict

def merger_data_consolidado(data_empleadores, data_patron, data_planilla):
    empleado = process_data_empleados(data_empleadores)
    patron = process_data_patron(data_patron)
    planilla = process_data_planilla(data_planilla)
    
    # Create a new dictionary to store the combined information grouped by NIT
    new_dictionary = {}

    # Combine the information from the 'employee' and 'employer' dictionaries
    for nit, datos in empleado.items():
        if nit not in new_dictionary:
            new_dictionary[nit] = {}
        new_dictionary[nit]['empleado'] = datos

    for nit, datos in patron.items():
        if nit not in new_dictionary:
            new_dictionary[nit] = {}
        new_dictionary[nit]['patron'] = datos      

    for nit, datos in planilla.items():
        if nit not in new_dictionary:
            new_dictionary[nit] = {}
        new_dictionary[nit]['planilla'] = datos      
    
    return new_dictionary

def save_data(sheet,data):
    merge_data = merger_data_consolidado(data['Valores empleado'], data['Valores patron'], data['Valores planilla'])

    # Variable to track the current row number in the excel sheet
    current_row = 3
    suma_empleado_patron = 0
    total_pago = 0
    current_tipo_entidad = None
    sum_subtotales = [0] * 13
    sum_totales = [0] * 13

    for nit, item in merge_data.items():
        entidad = Entidad.objects.filter(NIT=nit).first()
        suma_empleado_patron = convert_empty_to_zero(item.get('empleado', {}).get('TOTAL')) + convert_empty_to_zero(item.get('patron', {}).get('TOTAL'))

        if entidad.tipo != current_tipo_entidad:  
            if current_tipo_entidad is not None:
                add_subtotal(sheet, sum_subtotales, current_row, current_tipo_entidad)
                current_row += 1
           
            current_tipo_entidad = entidad.tipo
            for i in range(len(sum_subtotales)):
                if (i == 12):
                    sum_totales[i] = sum_subtotales[i]
                else:
                    sum_totales[i] += sum_subtotales[i]

            for i in range(12):
                sum_subtotales[i] = 0
        
        total_pago += suma_empleado_patron
        
        # Accumulate items
        sum_subtotales[0] += convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 2'))      
        sum_subtotales[1] += convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 8'))      
        sum_subtotales[2] += convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 9'))      
        sum_subtotales[3] += convert_empty_to_zero(item.get('empleado', {}).get('TOTAL'))
        sum_subtotales[4] += convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 2')) 
        sum_subtotales[5] += convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 8')) 
        sum_subtotales[6] += convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 9')) 
        sum_subtotales[7] += convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 2'))
        sum_subtotales[8] += convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 8'))
        sum_subtotales[9] += convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 9'))
        sum_subtotales[10] += convert_empty_to_zero(item.get('patron', {}).get('TOTAL'))
        sum_subtotales[11] = suma_empleado_patron
        sum_subtotales[12] += convert_empty_to_zero(item.get('planilla', {}).get('VALOR'))
    
        sheet[f"A{current_row}"] = entidad.NIT
        sheet[f"B{current_row}"] = entidad.rubro
        sheet[f"C{current_row}"] = entidad.concepto
        sheet[f"D{current_row}"] = convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 2')) 
        sheet[f"E{current_row}"] = convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 8'))        
        sheet[f"F{current_row}"] = convert_empty_to_zero(item.get('empleado', {}).get('UNIDAD 9')) 
        sheet[f"G{current_row}"] = convert_empty_to_zero(item.get('empleado', {}).get('TOTAL'))
        sheet[f"H{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 2')) 
        sheet[f"I{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 8')) 
        sheet[f"J{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('TEMPORAL UN 9'))
        sheet[f"K{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 2')) 
        sheet[f"L{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 8')) 
        sheet[f"M{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('PERMANENTE UN 9')) 
        sheet[f"N{current_row}"] = convert_empty_to_zero(item.get('patron', {}).get('TOTAL')) 
        sheet[f"O{current_row}"] = sum_subtotales[11]
        sheet[f"P{current_row}"] = convert_empty_to_zero(item.get('planilla', {}).get('VALOR'))

        # Add format styles 
        sheet[f"D{current_row}"].style = currency_style
        sheet[f"E{current_row}"].style = currency_style
        sheet[f"F{current_row}"].style = currency_style
        sheet[f"G{current_row}"].style = currency_style
        sheet[f"H{current_row}"].style = currency_style
        sheet[f"I{current_row}"].style = currency_style
        sheet[f"J{current_row}"].style = currency_style
        sheet[f"K{current_row}"].style = currency_style
        sheet[f"L{current_row}"].style = currency_style
        sheet[f"M{current_row}"].style = currency_style
        sheet[f"N{current_row}"].style = currency_style
        sheet[f"O{current_row}"].style = currency_style
        sheet[f"P{current_row}"].style = currency_style

        # Increment the current row number for the next iteration
        current_row += 1

        sum_subtotales[11] = total_pago

    if current_tipo_entidad is not None: 
        add_subtotal(sheet, sum_subtotales, current_row, current_tipo_entidad)
        current_row += 1
        print('cantidad',sum_subtotales)
        for i in range(len(sum_subtotales)):
                if (i == 12):
                    sum_totales[i] = sum_subtotales[i]
                else:
                    sum_totales[i] += sum_subtotales[i]

    add_totales(sheet, sum_totales, current_row)
    

def add_subtotal(sheet, sum_subtotales, current_row, tipo_entidad):      
    # Add the subtotal cells and apply the style
    sheet[f"B{current_row}"] = "Subtotal"
    sheet[f"C{current_row}"] = tipo_entidad
    sheet[f"B{current_row}"].font = bold_font
    sheet[f"C{current_row}"].font = bold_font
    for col_index, value in enumerate(sum_subtotales):
        col_letter = chr(ord('D') + col_index) 
        print(col_letter)
        if ( col_letter != 'P'):
            sheet[f"{col_letter}{current_row}"] = value
            sheet[f"{col_letter}{current_row}"].style = currency_style
            sheet[f"{col_letter}{current_row}"].font = bold_font
        else:
            continue

def add_totales(sheet, data, current_row):
    suma_empleado_patron = data[3] + data[10]

    total_data = [
        "",
        "A0102..",
        "TOTAL",
        data[0],
        data[1],
        data[2],
        data[3],
        data[4],
        data[5],
        data[6],
        data[7],
        data[8],
        data[9],
        data[10],
        suma_empleado_patron,
        data[12],
    ]

    additional_row_index = current_row 
    for col_idx, value in enumerate(total_data, start=1):
        cell = sheet.cell(row=additional_row_index, column=col_idx, value=value)
        if isinstance(value, (int, float)):
            cell.style = currency_style
        cell.font = bold_font


        