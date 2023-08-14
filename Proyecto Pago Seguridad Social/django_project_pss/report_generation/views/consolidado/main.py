import openpyxl 
from openpyxl import load_workbook

from django.http import HttpResponse
from document_upload.models import *
from ..functions import *
from ..constants import *
from .consolidado import *
from .resumen_temporales import *
from .resumen_permanentes import *

def get_data_values(date):
    try:
        empleados = valoresEmpleado.objects.filter(fecha=date).order_by('NIT__razonEntidad')
        empleados_ordenados = sorted(
            empleados, key=lambda empleado: PERSONALIZED_ORDER.index(empleado.NIT.razonEntidad)
        )

        patrones = valoresPatron.objects.filter(fecha=date).order_by('NIT__razonEntidad')
        patrones_ordenados = sorted(
            patrones, key=lambda patron: PERSONALIZED_ORDER.index(patron.NIT.razonEntidad)
        )

        # Create a dictionary to map entity names to indices
        entidad_to_index = {entidad: index for index, entidad in enumerate(PERSONALIZED_ORDER)}
        
        # Get the filtered objects and assign them order values
        valores_planilla_ordenados = valoresPlanilla.objects.filter(
            numeroPlanilla__fecha=date
        ).annotate(
            entidad_order=Case(
                *[When(codigoEntidad__concepto=entidad, then=Value(index)) for entidad, index in entidad_to_index.items()],
                default=Value(len(PERSONALIZED_ORDER)), output_field=CharField()
            )
        ).order_by('entidad_order')

        # Create a dictionary to store the information grouped by NIT
        data = {
            'Valores patron':patrones_ordenados,
            'Valores empleado':empleados_ordenados,
            'Valores planilla': valores_planilla_ordenados
        }
    
        return data
    
    except ObjectDoesNotExist:
        return None  # Return None if there's no data for the given date or any other related error
    except Exception as e:
        # Handle other exceptions as needed, such as database errors, unexpected behavior, etc.
        print(f"An error occurred: {e}")
        return None

def generate_consolidado_sheet(source_sheet, workbook, data):
    sheet_consolidado = workbook.active

    cells_range_entidad = "A1:C1"
    cells_range_empleado = "D1:G1"
    cells_range_patron = "H1:N1"

    # # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet_consolidado)
    
    sheet_consolidado.merge_cells(cells_range_entidad)
    sheet_consolidado.merge_cells(cells_range_empleado)
    sheet_consolidado.merge_cells(cells_range_patron)

    merged_ranges = [cells_range_entidad, cells_range_empleado, cells_range_patron]

    # Set alignment for merged cells
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(":")
        start_row, start_column = openpyxl.utils.coordinate_to_tuple(start_cell)
        end_row, end_column = openpyxl.utils.coordinate_to_tuple(end_cell)
        
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = sheet_consolidado.cell(row=row, column=column)
                cell.alignment = center_alignment
                # cell.font = bold_font

    # Assign the desired name to the sheet
    sheet_name = f"CONSOLIDADO"
    sheet_consolidado.title = sheet_name

    # Add additional information to the sheet
    save_data(sheet_consolidado, data)

def generate_masivo_temporales_sheet(source_sheet,workbook,date):
    sheet_archivos = workbook.create_sheet("RESUMEN MASIVO TEMPORALES")

    cells_range_1 = "A1:B1"
    cells_range_2 = "C1:D1"
    cells_range_3 = "F1:K1"

    # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet_archivos)

    sheet_archivos.merge_cells(cells_range_1)
    sheet_archivos.merge_cells(cells_range_2)
    sheet_archivos.merge_cells(cells_range_3)

    merged_ranges = [cells_range_1, cells_range_2, cells_range_3]

    # Set alignment for merged cells
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(":")
        start_row, start_column = openpyxl.utils.coordinate_to_tuple(start_cell)
        end_row, end_column = openpyxl.utils.coordinate_to_tuple(end_cell)
        
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = sheet_archivos.cell(row=row, column=column)
                cell.alignment = center_alignment
                # # cell.font = bold_font

    save_data_temporales(sheet_archivos,date)

def generate_masivo_permanentes_sheet(source_sheet,workbook,date):
    sheet_archivos = workbook.create_sheet("RESUMEN MASIVO PERMANENTES")

    cells_range_1 = "A1:B1"
    cells_range_2 = "C1:D1"
    cells_range_3 = "F1:K1"

    # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet_archivos)

    sheet_archivos.merge_cells(cells_range_1)
    sheet_archivos.merge_cells(cells_range_2)
    sheet_archivos.merge_cells(cells_range_3)

    merged_ranges = [cells_range_1, cells_range_2, cells_range_3]

    # Set alignment for merged cells
    for merged_range in merged_ranges:
        start_cell, end_cell = merged_range.split(":")
        start_row, start_column = openpyxl.utils.coordinate_to_tuple(start_cell)
        end_row, end_column = openpyxl.utils.coordinate_to_tuple(end_cell)
        
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                cell = sheet_archivos.cell(row=row, column=column)
                cell.alignment = center_alignment
                # # cell.font = bold_font

    save_data_permanentes(sheet_archivos,date)

def generate_excel_report(data, year, month):
    source_file_consolidado  = 'media/plantillas/Consolidado.xlsx'
    source_file_resumen_masivo = 'media/plantillas/Resumen_masivo.xlsx'

    date = f"{year}/{month}"

    # Load the existing excel file
    source_workbook_consolidado = load_workbook(source_file_consolidado)
    source_workbook_masivo_temporales = load_workbook(source_file_resumen_masivo)
    source_workbook_masivo_permanentes = load_workbook(source_file_resumen_masivo)

    # Get the first sheet of the existing excel file
    source_sheet_consolidado = source_workbook_consolidado.active
    source_sheet_masivo_temporales = source_workbook_masivo_temporales.active
    source_sheet_masivo_permanentes = source_workbook_masivo_permanentes.active

    # Create a new Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()

    generate_consolidado_sheet(source_sheet_consolidado, workbook, data)
    generate_masivo_temporales_sheet(source_sheet_masivo_temporales, workbook,date)     
    generate_masivo_permanentes_sheet(source_sheet_masivo_permanentes, workbook,date)     

    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte Final-{year}-{month}.xlsx"'
    workbook.save(response)

    return response