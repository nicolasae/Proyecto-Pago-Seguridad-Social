import openpyxl
from collections import defaultdict
from django.http import HttpResponse

from document_upload.models import *
from .constants import *
from .functions import *

def get_data_permanentes(date):
    orden_personalizado = [
        'SALUD',
        'RIESGOS PROFESIONALES',
        'PENSION',
        'MEN',
        'SENA',
        'ESAP',
        'ICBF',
        'CAJA DE COMPENSACION FAMILIAR',
    ]
    # Filtrar por tipoPatronal permanente y fecha
    motivos = valoresPatron.objects.filter(tipoPatronal__tipo='permanente', fecha=date).order_by('NIT__razonEntidad')
    motivos_ordenados = sorted(motivos, key=lambda motivo: orden_personalizado.index(motivo.NIT.razonEntidad))

    return motivos_ordenados

def generate_excel_report_permanentes(data, year, month):
    # Create an Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()
    sheet1 = create_sheet_permanentes(workbook)

    # Write the data for the first sheet
    write_permanentes_data(sheet1, data)

    # Assign the desired name to the sheet
    sheet_name = f"Datos"
    sheet1.title = sheet_name

    
    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Resumen Masivo Permanentes-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

def create_sheet_permanentes(workbook):

    sheet = workbook.active
    sheet.title = "Datos"

    headers_sheet = {
        1: "NIT",
        2: "RUBRO",
        3: "CONCEPTO",
        4: "CONDIGO DEL CONCEPTO DE DESCUENTO",
        5: "TIPO CUENTA POR PAGAR",
        6: "UNIDAD 2",
        7: "UNIDAD 8",
        8: "UNIDAD 9",
        9: "TOTAL",
    }

    for col_idx, header_text in headers_sheet.items():
        cell = sheet.cell(row=1, column=col_idx, value=header_text)

    return sheet


def write_permanentes_data(sheet, data):   
    suma_unidad2 = 0
    suma_unidad8 = 0
    suma_unidad9 = 0
    suma_total = 0

    for index, motivo in enumerate(data, start=2): 
        suma_unidad2 += motivo.unidad2
        suma_unidad8 += motivo.unidad8
        suma_unidad9 += motivo.unidad9
        suma_total += motivo.total

        sheet[f"A{index}"] = motivo.NIT_id
        sheet[f"B{index}"] = motivo.NIT.rubro
        sheet[f"C{index}"] = motivo.NIT.concepto
        sheet[f"D{index}"] = motivo.NIT.codigo  
        sheet[f"E{index}"] = motivo.NIT.tipoCuentaPagar  
        sheet[f"F{index}"] = motivo.unidad2
        sheet[f"G{index}"] = motivo.unidad8
        sheet[f"H{index}"] = motivo.unidad9
        sheet[f"I{index}"] = motivo.total

        # Add styles to the cells
        sheet[f"F{index}"].style = currency_style 
        sheet[f"G{index}"].style = currency_style 
        sheet[f"H{index}"].style = currency_style 
        sheet[f"I{index}"].style = currency_style 

    total_data = [
        "supernume",  
        "A0102..",
        "TOTAL",
        "",
        "",
        suma_unidad2,
        suma_unidad8,
        suma_unidad9,
        suma_total,
    ]

    additional_row_index = len(data) + 2 
    for col_idx, value in enumerate(total_data, start=1):
        cell = sheet.cell(row=additional_row_index, column=col_idx, value=value)
        if col_idx > 4:
            cell.style = currency_style
        cell.font = bold_font