import openpyxl

from django.http import HttpResponse
from django.db.models import Sum,Case, CharField, Value, When

from document_upload.models import *
from .functions import *
from .constants import *

# Map model fields to the headers in the Excel file
field_mapping = {
    'razonSocial': 'RAZON SOCIAL',
    'fecha': 'PERIODO',
    'identificacion': 'IDENTIFICACION',
    'codigoDependenciaSucursal': 'COD. DEPENDENCIA O SUCURSAL',
    'nomDependenciaSucursal': 'NOM. DEPENDENCIA O SUCURSAL',
    'fechaReporte': 'FECHA GENERACION REPORTE',
    'fechaLimitePago': 'FECHA LIMITE DE PAGO',
    'periodoPension': 'PERIODO PENSION',
    'periodoSalud': 'PERIODO SALUD',
    'numeroPlanilla': 'NUMERO PLANILLA',
    'totalCotizantes': 'TOTAL COTIZANTES',
    'PIN': 'REFERENCIA DE PAGO (PIN)',
    'tipoPlanilla': 'TIPO DE PLANILLA',
}

def get_info_planilla(date):   
    # Get the datasheet objects filtered by year and month
    info_planilla = infoPlanilla.objects.filter(fecha=date)
    return info_planilla

def get_values_planilla(date):    
    # Create a dictionary to map entity names to indices
    entidad_to_index = {entidad: index for index, entidad in enumerate(PERSONALIZED_ORDER)}
    
    # Get the filtered objects and assign them order values
    valores_planilla_filtrados = valoresPlanilla.objects.filter(
        numeroPlanilla__fecha=date
    ).annotate(
        entidad_order=Case(
            *[When(codigoEntidad__concepto=entidad, then=Value(index)) for entidad, index in entidad_to_index.items()],
            default=Value(len(PERSONALIZED_ORDER)), output_field=CharField()
        )
    ).order_by('entidad_order')

    return valores_planilla_filtrados

def generate_excel_report(info_planilla,values_planilla, year, month):
    # Create an Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()

    sheet1 = create_sheet_info_planilla(workbook)
    sheet2 = create_sheet_values_planilla(workbook)
    
    # Write the data to the first sheet
    write_info_planilla_data(sheet1, info_planilla)

    # Write the data to the second sheet
    write_values_planilla_data(sheet2, values_planilla)

    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Planilla_Detallada-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

def create_sheet_info_planilla(workbook):
    sheet = workbook.active
    sheet.title = f"Info Planilla"   

    for row_num, header in enumerate(field_mapping.values(), start=1):
        cell = sheet.cell(row=row_num, column=1, value=header)
        cell.font = bold_font

    return sheet

def create_sheet_values_planilla(workbook):
    sheet = workbook.active
    sheet = workbook.create_sheet(title='Valores Planilla')
    
    headers_sheet = {
        1: "CODIGO ENTIDAD",
        2: "NIT",
        3: "NOMBRE",
        4: "NUMERO AFILIADOS",
        5: "FONDO SOLIDARIDAD",
        6: "FONDO SUBSISTENCIA",
        7: "TOTAL INTERESES",
        8: "VALOR PAGAR SIN INTERESES",
        9: "VALOR PAGAR",
    }

    for col_idx, header_text in headers_sheet.items():
        cell = sheet.cell(row=1, column=col_idx, value=header_text)
        cell.font = bold_font

    return sheet

def write_info_planilla_data(sheet, info_planilla):   
    for row_num, header in enumerate(field_mapping.values(), start=1):
        sheet.cell(row=row_num, column=1, value=header)

    row_num = 1
    for obj in info_planilla:
        for key in field_mapping:
            attribute_name = key  # Get the attribute name from the field_mapping
            value = getattr(obj, attribute_name, '')  # Get the attribute value using getattr
            cell = sheet.cell(row = row_num, column = 2, value = value)
            cell.alignment = left_alignment
            row_num += 1

def write_values_planilla_data(sheet, values_planilla):
    distinct_razon_entidades = Entidad.objects.order_by('razonEntidad').values_list('razonEntidad', flat=True).distinct()
    
    distinct_razon_entidades_ordered = sorted(
        distinct_razon_entidades,
        key=lambda entidad: PERSONALIZED_ORDER.index(entidad) if entidad in PERSONALIZED_ORDER else len(PERSONALIZED_ORDER)
    )

    # Initialize row_index for the first group
    row_index = 2

    for razon_entidad in distinct_razon_entidades_ordered:
        # Filter values_planilla for the current razon_entidad
        values_planilla_filtered = values_planilla.filter(codigoEntidad__razonEntidad=razon_entidad)

        suma_num_afiliados = 0
        suma_fondo_solidaridad = 0
        suma_fondo_subsistencia = 0
        suma_intereses = 0
        suma_sin_intereses = 0
        suma_total = 0

        for planilla in values_planilla_filtered: 
            suma_num_afiliados += planilla.numeroAfiliados
            suma_fondo_solidaridad += planilla.fondoSolidaridad
            suma_fondo_subsistencia += planilla.fondoSubsistencia
            suma_intereses += planilla.totalIntereses
            suma_sin_intereses += planilla.valorPagarSinIntereses
            suma_total += planilla.valorPagar

            sheet[f"A{row_index}"] = planilla.codigoEntidad.codigo
            sheet[f"B{row_index}"] = planilla.NIT
            sheet[f"C{row_index}"] = planilla.codigoEntidad.concepto
            sheet[f"D{row_index}"] = planilla.numeroAfiliados
            sheet[f"E{row_index}"] = planilla.fondoSolidaridad
            sheet[f"F{row_index}"] = planilla.fondoSubsistencia
            sheet[f"G{row_index}"] = planilla.totalIntereses
            sheet[f"H{row_index}"] = planilla.valorPagarSinIntereses
            sheet[f"I{row_index}"] = planilla.valorPagar

            sheet[f"E{row_index}"].style = currency_style
            sheet[f"F{row_index}"].style = currency_style 
            sheet[f"G{row_index}"].style = currency_style
            sheet[f"H{row_index}"].style = currency_style 
            sheet[f"I{row_index}"].style = currency_style

            # Increment the row_index for the next row
            row_index += 1

        total_data = [
            "",
            "",
            f"SUBTOTAL {razon_entidad}",
            suma_num_afiliados,
            suma_fondo_solidaridad,
            suma_fondo_subsistencia,
            suma_intereses,
            suma_sin_intereses,
            suma_total
        ]

        sheet.cell(row=row_index, column=1, value=total_data[0])
        sheet.cell(row=row_index, column=2, value=total_data[1])
        cell = sheet.cell(row=row_index, column=3, value=total_data[2])
        cell.font = bold_font

        for col_idx, value in enumerate(total_data[3:], start=4):
            cell = sheet.cell(row=row_index, column=col_idx, value=value)
            if col_idx > 4:
                cell.style = currency_style
            cell.font = bold_font

        # Increment the row_index for the next group
        row_index += 1

    # Calculate the grand total using all values in values_planilla
    grand_total_data = [
        "",
        "",
        "TOTAL",
        values_planilla.aggregate(Sum('numeroAfiliados'))['numeroAfiliados__sum'],
        values_planilla.aggregate(Sum('fondoSolidaridad'))['fondoSolidaridad__sum'],
        values_planilla.aggregate(Sum('fondoSubsistencia'))['fondoSubsistencia__sum'],
        values_planilla.aggregate(Sum('totalIntereses'))['totalIntereses__sum'],
        values_planilla.aggregate(Sum('valorPagarSinIntereses'))['valorPagarSinIntereses__sum'],
        values_planilla.aggregate(Sum('valorPagar'))['valorPagar__sum']
    ]

    sheet.cell(row=row_index, column=1, value=grand_total_data[0])
    sheet.cell(row=row_index, column=2, value=grand_total_data[1])
    cell = sheet.cell(row=row_index, column=3, value=grand_total_data[2])
    cell.font = bold_font
    
    for col_idx, value in enumerate(grand_total_data[3:], start=4):
        cell = sheet.cell(row=row_index, column=col_idx, value=value)
        if col_idx > 4:
            cell.style = currency_style
        cell.font = bold_font
        