import openpyxl
from django.http import HttpResponse

from document_upload.models import *

def get_data_temporales(date):
    orden_personalizado = [
        'SALUD',
        'RIESGOS PROFESIONALES',
        'PENSION',
        'MEN',
        'SENA',
        'ESAP',
        'ICBF',
        'CAJA DE COMPENSACION FAMILIAR',
    ]
    # Filtrar por tipoPatronal permanente y fecha
    motivos = Motivo.objects.filter(tipoPatronal__tipo='temporal', fecha=date).order_by('NIT__razonEntidad')
    motivos_ordenados = sorted(motivos, key=lambda motivo: orden_personalizado.index(motivo.NIT.razonEntidad))

    return motivos_ordenados

def generate_excel_report_temporales(data, year, month):
    # Crear un archivo de Excel para los datos de la planilla
    workbook = openpyxl.Workbook()
    sheet1 = create_sheet_temporales(workbook)

    # Escribir los datos para la primera hoja
    write_temporales_data(sheet1, data)

    # Asignar el nombre deseado a la hoja
    sheet_name = f"Datos"
    sheet1.title = sheet_name

    # Crear la respuesta HTTP y devolver el archivo para su descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Resumen Masivo Temporales-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

def create_sheet_temporales(workbook):

    sheet = workbook.active
    sheet.title = "Temporales Data"

    headers_sheet = {
        1: "NIT",
        2: "RUBRO",
        3: "CONCEPTO",
        4: "CODIGO DEL CONCEPTO DE DESCUENTO",
        5: "TIPO CUENTA POR PAGAR",
        6: "TEMPORAL 2",
        7: "TEMPORAL 8",
        8: "TEMPORAL 9",
        9: "TOTAL",
    }

    for col_idx, header_text in headers_sheet.items():
        cell = sheet.cell(row=1, column=col_idx, value=header_text)

    return sheet


def write_temporales_data(sheet, data):   
    suma_unidad2 = 0
    suma_unidad8 = 0
    suma_unidad9 = 0
    suma_total = 0

    for index, motivo in enumerate(data, start=2): 
        suma_unidad2 += motivo.unidad2
        suma_unidad8 += motivo.unidad8
        suma_unidad9 += motivo.unidad9
        suma_total += motivo.total

        sheet[f"A{index}"] = motivo.NIT_id
        sheet[f"B{index}"] = motivo.NIT.rubro
        sheet[f"C{index}"] = motivo.NIT.concepto
        sheet[f"D{index}"] = motivo.NIT.codigo  
        sheet[f"E{index}"] = motivo.NIT.tipoCuentaPagar  
        sheet[f"F{index}"] = motivo.unidad2
        sheet[f"G{index}"] = motivo.unidad8
        sheet[f"H{index}"] = motivo.unidad9
        sheet[f"I{index}"] = motivo.total


    total_data = [
        "supernume",  
        "A0102..",
        "TOTAL",
        "",
        "",
        suma_unidad2,
        suma_unidad8,
        suma_unidad9,
        suma_total,
    ]

    additional_row_index = len(data) + 2 
    for col_idx, value in enumerate(total_data, start=1):
        sheet.cell(row=additional_row_index, column=col_idx, value=value)