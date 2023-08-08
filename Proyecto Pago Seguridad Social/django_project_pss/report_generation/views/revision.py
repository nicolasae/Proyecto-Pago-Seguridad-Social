import openpyxl
from django.http import HttpResponse
from django.db.models import Sum

from document_upload.models import *

def revision(request,year, month):
    
    return create_excel(year, month)
    
def get_data(year,month):
    date = year + '/' + month
    value_planilla = valoresPlanilla.objects.filter(numeroPlanilla__periodo=date).aggregate(Sum('valorPagar'))['valorPagar__sum']
    value_temporal = valoresPatron.objects.filter(tipoPatronal__tipo='temporal',fecha=date).aggregate(Sum('total'))['total__sum']
    value_permanente = valoresPatron.objects.filter(tipoPatronal__tipo='permanente',fecha=date).aggregate(Sum('total'))['total__sum']
    value_patron = value_permanente + value_temporal
    empleado_un2 = valoresEmpleado.objects.filter(unidad=2,periodo=date).aggregate(Sum('saldo'))['saldo__sum']
    empleado_un8 = valoresEmpleado.objects.filter(unidad=8,periodo=date).aggregate(Sum('saldo'))['saldo__sum']
    empleado_un9 = valoresEmpleado.objects.filter(unidad=9,periodo=date).aggregate(Sum('saldo'))['saldo__sum']
    value_empleado = empleado_un2 + empleado_un8 + empleado_un2
    total = value_patron + value_empleado

    values = {
        'planilla':value_planilla,
        'temporal':value_temporal,
        'permanente':value_permanente,
        'patron':value_patron,
        'empleado2':empleado_un2,
        'empleado8':empleado_un8,
        'empleado9':empleado_un9,
        'empleado': value_empleado,
        'total': total,
    }

    print(values)

def create_excel(year,month):  
    workbook = openpyxl.Workbook()
    sheet1 = create_sheet(workbook)
    
    # Write data
    get_data(year,month)
    # write_revision_data(sheet1, data)

    # Crear la respuesta HTTP y devolver el archivo para su descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Revision-{year}-{month}.xlsx"'
    workbook.save(response)

    return response

def create_sheet(workbook):
    sheet = workbook.active
    sheet.title = f"Revision"

    # Mapeo de campos del modelo con los encabezados en el archivo Excel
    field_mapping = {
        'temporal': 'Aporte Patronal Planta Temporal',
        'permanente': 'Aporte Patronal Planta Permanente',
        'empleadoUnidad2':'Empleado Unidad 2',
        'empleadoUnidad8':'Empleado Unidad 8',
        'empleadoUnidad9':'Empleado Unidad 9',
        'total':'Total',
        'planilla':'Planilla',
        'diferencia':'Diferencia'
    }

    # Escribir los encabezados en la columna A
    for row_num, header in enumerate(field_mapping.values(), start=1):
        sheet.cell(row=row_num, column=1, value=header)

    return sheet

# def write_revision_data(sheet,data):