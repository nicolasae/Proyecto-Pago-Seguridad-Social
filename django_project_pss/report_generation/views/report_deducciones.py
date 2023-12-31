import openpyxl 
from openpyxl import load_workbook

from django.http import HttpResponse
from document_upload.models import *
from .functions import * 
from .constants import * 

def get_data_deducciones(date):
    motivos = valoresEmpleado.objects.filter(fecha=date).order_by('NIT__razonEntidad')
    motivos_ordenados = sorted(motivos, key=lambda motivo: PERSONALIZED_ORDER.index(motivo.NIT.razonEntidad))

    # Create a dictionary to store the information grouped by NIT
    data_dict = {}

    for motivo in motivos_ordenados:
        NIT = motivo.NIT.NIT
        rubroTemporal = motivo.NIT.rubroTemporal
        rubroPermanente = motivo.NIT.rubroPermanente
        concepto = motivo.NIT.concepto
        unidad = motivo.unidad

        # # Check if an entry for the NIT already exists in the dictionary
        if NIT not in data_dict:
            data_dict[NIT] = {
                'RUBRO TEMPORAL': rubroTemporal,
                'RUBRO PERMANENTE': rubroPermanente,
                'CONCEPTO': concepto,
                'UNIDAD 2': 0,
                'UNIDAD 8': 0,
                'UNIDAD 9': 0,
                'TOTAL': 0
            }

        # # Add the information to the corresponding entry of type_employer
        data_dict[NIT]['TOTAL'] += motivo.saldo

        if unidad == 2:
            data_dict[NIT]['UNIDAD 2'] += motivo.saldo
        elif unidad == 8:
            data_dict[NIT]['UNIDAD 8'] += motivo.saldo
        elif unidad == 9:
            data_dict[NIT]['UNIDAD 9'] += motivo.saldo
        
    return data_dict

def save_data_deducciones(sheet,data):
    suma_unidad2 = 0
    suma_unidad8 = 0
    suma_unidad9 = 0
    suma_total = 0
    # Variable to track the current row number in the excel sheet
    current_row = 2

    # Loop through the dictionary for each NIT
    for nit, item in data.items():
        suma_unidad2 += item['UNIDAD 2']
        suma_unidad8 += item['UNIDAD 8']
        suma_unidad9 += item['UNIDAD 9']
        
        sheet[f"A{current_row}"] = nit
        sheet[f"B{current_row}"] = item['RUBRO TEMPORAL']
        sheet[f"C{current_row}"] = item['RUBRO PERMANENTE']
        sheet[f"D{current_row}"] = item['CONCEPTO']
        sheet[f"E{current_row}"] = item['UNIDAD 2']
        sheet[f"F{current_row}"] = item['UNIDAD 8']
        sheet[f"G{current_row}"] = item['UNIDAD 9']
        sheet[f"H{current_row}"] = item['TOTAL']
        
        # Add format styles 
        columns_to_style = ['A', 'B', 'C', 'D', 'E', 'F', 'G','H']

        for col in columns_to_style:
            cell = f"{col}{current_row}"

            if col in ['E', 'F','G','H']:
                sheet[cell].style = currency_style 
            sheet[cell].font = font_style
            sheet[cell].border = border_style

        # Increment the current row number for the next iteration
        current_row += 1

    suma_total += suma_unidad2 + suma_unidad8 + suma_unidad9
    total_data = [
        "",  
        "A0102..",
        "A0101..",
        "TOTAL",
        suma_unidad2,
        suma_unidad8,
        suma_unidad9,       
        suma_total,
    ]

    additional_row_index = len(data) + 2 
    for col_idx, value in enumerate(total_data, start=1):
        cell = sheet.cell(row=additional_row_index, column=col_idx, value=value)
        if col_idx > 3:
            cell.style = currency_style
        cell.font = header_style
        cell.border = border_style

def generate_excel_report_deducciones(data, year, month):
    source_file  = 'media/plantillas/Resumen_deducciones.xlsx'
    # Load the existing excel file
    source_workbook = load_workbook(source_file)

    # Get the first sheet of the existing excel file
    source_sheet = source_workbook.active

    # Create a new Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active

    # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet1)

    # Assign the desired name to the sheet
    sheet_name = f"Deducciones-{year}-{month}"
    sheet1.title = sheet_name

    # Add additional information to the sheet
    save_data_deducciones(sheet1, data)

    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Resumen Deducciones-{year}-{month}.xlsx"'
    workbook.save(response)

    return response