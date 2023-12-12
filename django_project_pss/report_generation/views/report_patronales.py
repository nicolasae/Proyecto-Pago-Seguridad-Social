import openpyxl 
from openpyxl import load_workbook

from django.http import HttpResponse
from document_upload.models import *
from .functions import *
from .constants import *

def get_data_patronales(date):    
    motivos = valoresPatron.objects.filter(fecha=date).order_by('NIT__razonEntidad')
    motivos_ordenados = sorted(motivos, key=lambda motivo: PERSONALIZED_ORDER.index(motivo.NIT.razonEntidad))

    # Create a dictionary to store the information grouped by NIT
    data_dict = {}

    for motivo in motivos_ordenados:
        NIT = motivo.NIT.NIT
        tipo_patronal = motivo.tipoPatronal.tipo
        rubroPermanente = motivo.NIT.rubroPermanente
        rubroTemporal = motivo.NIT.rubroTemporal
        concepto = motivo.NIT.concepto
        codigo_descuento = motivo.NIT.codigoDescuento
        total = motivo.total

        # Check if an entry for the NIT already exists in the dictionary
        if NIT not in data_dict:
            data_dict[NIT] = {
                'RUBRO PERMANENTE': rubroPermanente,
                'RUBRO TEMPORAL': rubroTemporal,
                'CONCEPTO': concepto,
                'CODIGO DEL CONCEPTO DE DESCUENTO': codigo_descuento,
                'TEMPORAL UN 2': 0,
                'TEMPORAL UN 8': 0,
                'TEMPORAL UN 9': 0,
                'PERMANENTE UN 2': 0,
                'PERMANENTE UN 8': 0,
                'PERMANENTE UN 9': 0,
                'TOTAL': 0
            }

        # Add the information to the corresponding entry of type_employer
        if tipo_patronal == 'temporal':
            data_dict[NIT]['TEMPORAL UN 2'] += motivo.unidad2
            data_dict[NIT]['TEMPORAL UN 8'] += motivo.unidad8
            data_dict[NIT]['TEMPORAL UN 9'] += motivo.unidad9
        elif tipo_patronal == 'permanente':
            data_dict[NIT]['PERMANENTE UN 2'] += motivo.unidad2
            data_dict[NIT]['PERMANENTE UN 8'] += motivo.unidad8
            data_dict[NIT]['PERMANENTE UN 9'] += motivo.unidad9

        # We also store the information in the entry of the employer type
        data_dict[NIT][tipo_patronal] = data_dict[NIT].get(tipo_patronal, 0) + motivo.unidad2

        data_dict[NIT]['TOTAL'] += total

    return data_dict

def save_data_patronales(sheet,data):
    suma_temporal2 = 0
    suma_temporal8 = 0
    suma_temporal9 = 0
    suma_permanente2 = 0
    suma_permanente8 = 0
    suma_permanente9 = 0
    suma_total = 0
    # Variable to track the current row number in the excel sheet
    current_row = 2

    # Loop through the dictionary for each NIT
    for nit, item in data.items():
        suma_temporal2 += item['TEMPORAL UN 2']
        suma_temporal8 += item['TEMPORAL UN 8']
        suma_temporal9 += item['TEMPORAL UN 9']
        suma_permanente2 += item['PERMANENTE UN 2']
        suma_permanente8 += item['PERMANENTE UN 8']
        suma_permanente9 += item['PERMANENTE UN 9']
        suma_total += item['TOTAL']

        sheet[f"A{current_row}"] = nit
        sheet[f"B{current_row}"] = item['RUBRO TEMPORAL']
        sheet[f"C{current_row}"] = item['RUBRO PERMANENTE']
        sheet[f"D{current_row}"] = item['CONCEPTO']
        sheet[f"E{current_row}"] = item['CODIGO DEL CONCEPTO DE DESCUENTO']
        sheet[f"F{current_row}"] = item['TEMPORAL UN 2']
        sheet[f"G{current_row}"] = item['TEMPORAL UN 8']
        sheet[f"H{current_row}"] = item['TEMPORAL UN 9']
        sheet[f"I{current_row}"] = item['PERMANENTE UN 2']
        sheet[f"J{current_row}"] = item['PERMANENTE UN 8']
        sheet[f"K{current_row}"] = item['PERMANENTE UN 9']
        sheet[f"L{current_row}"] = item['TOTAL']

        # Add format styles 
        columns_to_style = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I','J','K','L']

        for col in columns_to_style:
            cell = f"{col}{current_row}"

            if col in ['F','G', 'H','I','J','K','L']:
                sheet[cell].style = currency_style 
            sheet[cell].font = font_style
            sheet[cell].border = border_style
               
        # Increment the current row number for the next iteration
        current_row += 1

    total_data = [
        "",  
        "A0102..",
        "A0101..",
        "",
        "TOTAL",
        suma_temporal2,
        suma_temporal8,
        suma_temporal9,
        suma_permanente2,
        suma_permanente8,
        suma_permanente9,        
        suma_total,
    ]

    additional_row_index = len(data) + 2 
    for col_idx, value in enumerate(total_data, start=1):
        cell = sheet.cell(row=additional_row_index, column=col_idx, value=value)
        if col_idx > 1:
            cell.style = currency_style
        cell.font = header_style
        cell.border = border_style

def generate_excel_report_patronales(data, year, month):

    source_file  = 'media/plantillas/Resumen_patronales.xlsx'
    # Load the existing excel file
    source_workbook = load_workbook(source_file)

    # Get the first sheet of the existing excel file
    source_sheet = source_workbook.active

    # Create a new Excel file for the spreadsheet data
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active

    # Copy the data from the source sheet to the new sheet
    copy_data_from_existing_sheet(source_sheet, sheet1)

    # Assign the desired name to the sheet
    sheet_name = f"Datos-{year}-{month}"
    sheet1.title = sheet_name

    # Add additional information to the sheet
    save_data_patronales(sheet1, data)

    # Create the HTTP response and return the file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Resumen Patronales-{year}-{month}.xlsx"'
    workbook.save(response)

    return response